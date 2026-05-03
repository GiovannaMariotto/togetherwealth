from __future__ import annotations

from io import BytesIO
import pandas as pd
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from functions.analytics import filter_dashboard_transactions, prepare_cashflow_chart_data, prepare_summary, to_numeric_amounts


TRANSACTION_COLUMNS = [
    "entry_date",
    "month",
    "partner",
    "transaction_type",
    "category",
    "subcategory",
    "source",
    "amount",
    "notes",
]


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for col in ws.columns:
        max_len = 12
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 35)


def _safe_sheet_name(name: str) -> str:
    invalid_chars = "[]:*?/\\"
    cleaned = "".join("-" if char in invalid_chars else char for char in name)
    return cleaned[:31]


def _normalize_transactions(transactions: pd.DataFrame) -> pd.DataFrame:
    df = transactions.copy()
    if df.empty:
        return pd.DataFrame(columns=TRANSACTION_COLUMNS)

    for column in TRANSACTION_COLUMNS:
        if column not in df.columns:
            df[column] = None
    df = df[TRANSACTION_COLUMNS]
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)
    df["transaction_type"] = df["transaction_type"].fillna("").astype(str)
    return df


def _dashboard_table(df: pd.DataFrame, view_name: str) -> pd.DataFrame:
    summary = prepare_summary(df)
    _, net_cashflow = prepare_cashflow_chart_data(df)
    return pd.DataFrame(
        [
            ["View", view_name],
            ["Total Income", summary["income"]],
            ["Total Expenses", summary["expenses"]],
            ["Total Savings", summary["savings"]],
            ["Total Investments", summary["investments"]],
            ["Income After Expenses", net_cashflow],
            ["Net Worth Proxy", summary["net_worth_proxy"]],
            ["Savings Rate", summary["savings_rate"] / 100],
        ],
        columns=["Metric", "Value"],
    )


def _write_dashboard_sheet(writer, df: pd.DataFrame, sheet_name: str, view_name: str) -> None:
    dashboard = _dashboard_table(df, view_name)
    dashboard.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]

    cashflow, _ = prepare_cashflow_chart_data(df)
    if not cashflow.empty:
        start_row = 2
        ws.cell(start_row, 4, "Staggered Expenses")
        ws.cell(start_row + 1, 4, "Voice")
        ws.cell(start_row + 1, 5, "Amount")
        ws.cell(start_row + 1, 6, "Type")
        for idx, row in cashflow.iterrows():
            ws.cell(start_row + idx + 2, 4, row["voice"])
            ws.cell(start_row + idx + 2, 5, row["amount"])
            ws.cell(start_row + idx + 2, 6, row["measure"])

        chart = BarChart()
        chart.title = "Staggered Expenses"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Voice"
        data = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + len(cashflow) + 1)
        labels = Reference(ws, min_col=4, min_row=start_row + 2, max_row=start_row + len(cashflow) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "H2")

    expenses = df[df["transaction_type"] == "Expense"]
    if not expenses.empty:
        breakdown = expenses.groupby("category", as_index=False)["amount"].sum()
        start_row = 16
        ws.cell(start_row, 4, "Expense Breakdown")
        ws.cell(start_row + 1, 4, "Category")
        ws.cell(start_row + 1, 5, "Amount")
        for idx, row in breakdown.iterrows():
            ws.cell(start_row + idx + 2, 4, row["category"])
            ws.cell(start_row + idx + 2, 5, row["amount"])
        chart = PieChart()
        labels = Reference(ws, min_col=4, min_row=start_row + 2, max_row=start_row + len(breakdown) + 1)
        data = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + len(breakdown) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Expenses by Category"
        ws.add_chart(chart, "H18")


def build_excel_report(
    transactions: pd.DataFrame,
    projections: pd.DataFrame | None = None,
    dashboard_owners: list[str] | None = None,
) -> bytes:
    output = BytesIO()
    df = _normalize_transactions(transactions)
    dashboard_owners = dashboard_owners or ["All"]

    income = df[df["transaction_type"] == "Income"]
    expenses = df[df["transaction_type"] == "Expense"]
    savings = df[df["transaction_type"] == "Saving"]
    investments = df[df["transaction_type"] == "Investment"]
    if projections is None or projections.empty:
        projections = pd.DataFrame(columns=["Year", "Projected Value", "Monthly Contribution", "Annual Return %", "EURIBOR % Used"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for owner in dashboard_owners:
            view_df = filter_dashboard_transactions(df, "All months", owner)
            sheet_name = "Dashboard" if owner == "All" else _safe_sheet_name(f"Dashboard {owner}")
            _write_dashboard_sheet(writer, view_df, sheet_name, owner)

        df.to_excel(writer, sheet_name="All Inputs", index=False)
        income.to_excel(writer, sheet_name="Income", index=False)
        expenses.to_excel(writer, sheet_name="Expenses", index=False)
        savings.to_excel(writer, sheet_name="Savings", index=False)
        investments.to_excel(writer, sheet_name="Investments", index=False)
        projections.to_excel(writer, sheet_name="Projections", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            _style_sheet(ws)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        metric = str(ws.cell(row=cell.row, column=1).value)
                        if cell.column_letter == "B" and metric == "Savings Rate":
                            cell.number_format = "0.00%"
                        elif "rate" in str(ws.cell(row=1, column=cell.column).value).lower() or "%" in str(ws.cell(row=1, column=cell.column).value):
                            cell.number_format = "0.00"
                        else:
                            cell.number_format = '€#,##0.00'
    output.seek(0)
    return output.read()
